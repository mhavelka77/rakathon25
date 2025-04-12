import os
import tempfile
from typing import List
from fastapi import UploadFile
import docx
from pdf2image import convert_from_path
import pytesseract
from PIL import Image
import openpyxl

async def process_documents(files: List[UploadFile]) -> List[str]:
    results = []
    
    for file in files:
        file_content = await file.read()
        file_extension = os.path.splitext(file.filename)[1].lower()
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as temp_file:
            temp_file.write(file_content)
            temp_file_path = temp_file.name
        
        try:
            if file_extension == '.pdf':
                text = extract_text_from_pdf(temp_file_path)
            elif file_extension == '.docx':
                text = extract_text_from_docx(temp_file_path)
            elif file_extension in ['.jpg', '.jpeg', '.png']:
                text = extract_text_from_image(temp_file_path)
            elif file_extension == '.xlsx':
                text = extract_text_from_xlsx(temp_file_path)
            elif file_extension == '.txt':
                text = file_content.decode('utf-8')
            else:
                text = f"Unsupported file format: {file_extension}"
            
            results.append(text)
        
        finally:
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
        
        await file.seek(0)
    
    return results

def extract_text_from_pdf(file_path: str) -> str:
    try:
        images = convert_from_path(file_path)
        text_content = [pytesseract.image_to_string(image) for image in images]
        return "\n\n".join(text_content)
    except Exception as e:
        return f"Error extracting text from PDF: {str(e)}"

def extract_text_from_docx(file_path: str) -> str:
    try:
        doc = docx.Document(file_path)
        return "\n".join([paragraph.text for paragraph in doc.paragraphs])
    except Exception as e:
        return f"Error extracting text from DOCX: {str(e)}"

def extract_text_from_image(file_path: str) -> str:
    try:
        image = Image.open(file_path)
        return pytesseract.image_to_string(image)
    except Exception as e:
        return f"Error extracting text from image: {str(e)}"

def extract_text_from_xlsx(file_path: str) -> str:
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text_content = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text = [f"Sheet: {sheet_name}"]
            
            for row in sheet.iter_rows(values_only=True):
                row_text = [str(cell) if cell is not None else "" for cell in row]
                sheet_text.append(",".join(row_text))
            
            text_content.append("\n".join(sheet_text))
        
        return "\n\n".join(text_content)
    except Exception as e:
        return f"Error extracting text from XLSX: {str(e)}"